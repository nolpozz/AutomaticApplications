"""End-to-end pipeline and Excel-sync integration tests."""

from __future__ import annotations

import pytest

from job_agent.database.base import Database
from job_agent.database.repository import Repository
from job_agent.excel.sync import ExcelSynchronizer
from job_agent.models.domain import JobState
from job_agent.orchestrator.pipeline import Pipeline


@pytest.mark.integration
def test_full_pipeline_reaches_review(settings) -> None:  # type: ignore[no-untyped-def]
    pipeline = Pipeline(settings, formats=["md"])
    report = pipeline.run(offline=True)
    assert report.new > 0
    assert report.parsed == report.new
    assert report.ready_for_review > 0
    assert not report.errors

    counts = pipeline.stats().by_state
    assert counts.get(JobState.READY_FOR_REVIEW.value, 0) == report.ready_for_review


@pytest.mark.integration
def test_pipeline_is_idempotent(settings) -> None:  # type: ignore[no-untyped-def]
    pipeline = Pipeline(settings, formats=["md"])
    first = pipeline.run(offline=True)
    second = pipeline.run(offline=True)
    assert second.new == 0
    assert second.duplicates == first.scraped
    assert second.resumes == 0


@pytest.mark.integration
def test_excel_sync_creates_all_sheets(settings) -> None:  # type: ignore[no-untyped-def]
    from openpyxl import load_workbook

    Pipeline(settings, formats=["md"]).run(offline=True)
    db = Database(settings.storage.sqlite_path)
    with db.session_scope() as session:
        path = ExcelSynchronizer(Repository(session), settings.storage.excel_path).sync()
    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Jobs", "Applications", "Companies", "Documents", "Statistics"}
    jobs = wb["Jobs"]
    assert jobs.freeze_panes == "A2"
    assert jobs.max_row > 1


@pytest.mark.integration
def test_top_per_company_deprioritizes_extra_roles(settings) -> None:  # type: ignore[no-untyped-def]
    settings.pipeline.top_per_company = 2  # sample companies have 3 roles each
    pipeline = Pipeline(settings, formats=["md"])
    report = pipeline.run(offline=True)
    assert report.deprioritized > 0
    counts = pipeline.stats().by_state
    assert counts.get(JobState.DEPRIORITIZED.value, 0) == report.deprioritized
    # Deprioritized jobs never reach the expensive stages.
    assert report.parsed == report.new - report.deprioritized


@pytest.mark.integration
def test_top_per_company_zero_disables_cap(settings) -> None:  # type: ignore[no-untyped-def]
    settings.pipeline.top_per_company = 0
    report = Pipeline(settings, formats=["md"]).run(offline=True)
    assert report.deprioritized == 0
    assert report.parsed == report.new


def test_per_stage_models_resolve() -> None:
    from job_agent.config.settings import LLMSettings

    cfg = LLMSettings(model="gpt-4o", parse_model="gpt-4o-mini", classify_model="gpt-4o-mini")
    assert cfg.model_for("parse") == "gpt-4o-mini"
    assert cfg.model_for("classify") == "gpt-4o-mini"
    assert cfg.model_for("resume") == "gpt-4o"  # falls back to default
    assert cfg.model_for("cover_letter") == "gpt-4o"
