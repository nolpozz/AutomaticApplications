"""Excel synchronization — SQLite is the source of truth, Excel is a projection.

The workbook is regenerated from the database on demand (and automatically after
pipeline modifications when ``pipeline.auto_sync_excel`` is set). Excel is never
read back into the system. Sheets: Jobs, Applications, Companies, Documents,
Statistics — with row colors by stage, auto-sized columns, frozen headers, and
autofilter.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from job_agent.config.logging import get_logger
from job_agent.database.models import Company, Document
from job_agent.database.repository import Repository
from job_agent.models.domain import JobState

logger = get_logger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Row background by pipeline stage.
_STATE_FILL = {
    JobState.DISCOVERED.value: "F3F4F6",
    JobState.EMBEDDED.value: "E5E7EB",
    JobState.DEPRIORITIZED.value: "F1F5F9",
    JobState.PARSED.value: "E5E7EB",
    JobState.CLASSIFIED.value: "DBEAFE",
    JobState.REJECTED.value: "Fee2E2",
    JobState.READY_FOR_RESUME.value: "FEF3C7",
    JobState.RESUME_GENERATED.value: "FEF3C7",
    JobState.COVER_LETTER_GENERATED.value: "FDE68A",
    JobState.READY_FOR_REVIEW.value: "FDE68A",
    JobState.APPROVED.value: "D1FAE5",
    JobState.SUBMITTED.value: "A7F3D0",
    JobState.REJECTED_BY_COMPANY.value: "FECACA",
    JobState.INTERVIEW.value: "BFDBFE",
    JobState.OFFER.value: "6EE7B7",
    JobState.ARCHIVED.value: "E5E7EB",
}

_JOBS_HEADERS = [
    "Company",
    "Position",
    "Location",
    "Remote",
    "Salary",
    "URL",
    "Source",
    "Date Found",
    "Relevance",
    "Classification Score",
    "Recommendation",
    "Current Stage",
    "Application Status",
    "Resume Version",
    "Cover Letter Version",
    "Notes",
]


class ExcelSynchronizer:
    def __init__(self, repository: Repository, out_path: Path | str) -> None:
        self.repo = repository
        self.out_path = Path(out_path)

    def sync(self) -> Path:
        wb = Workbook()
        wb.remove(wb.active)  # drop default sheet
        self._build_jobs(wb.create_sheet("Jobs"))
        self._build_applications(wb.create_sheet("Applications"))
        self._build_companies(wb.create_sheet("Companies"))
        self._build_documents(wb.create_sheet("Documents"))
        self._build_statistics(wb.create_sheet("Statistics"))
        self.out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.out_path)
        logger.info("Synced Excel workbook -> %s", self.out_path)
        return self.out_path

    # -- sheets -------------------------------------------------------------
    def _build_jobs(self, ws: Worksheet) -> None:
        self._write_header(ws, _JOBS_HEADERS)
        jobs = self.repo.list_jobs()
        for job in jobs:
            score = self.repo.get_classifier(job.id)
            app = job.application
            resume = self.repo.latest_resume(job.id)
            cover = self.repo.latest_cover_letter(job.id)
            row = [
                job.company_name,
                job.title,
                job.location or "",
                job.remote,
                job.salary or "",
                job.url,
                job.source,
                _fmt_dt(job.created_at),
                (job.raw or {}).get("relevance", ""),
                round(score.overall_score, 3) if score else "",
                score.recommendation.value if score else "",
                job.state,
                app.status if app else "",
                f"v{resume.version}" if resume else "",
                f"v{cover.version}" if cover else "",
                (app.notes if app and app.notes else "") or (job.error or ""),
            ]
            ws.append(row)
            self._color_row(ws, ws.max_row, job.state, len(row))
        _finalize(ws)

    def _build_applications(self, ws: Worksheet) -> None:
        headers = ["Company", "Position", "Status", "Stage", "Submitted", "Responded", "Notes"]
        self._write_header(ws, headers)
        for app in self.repo.list_applications():
            job = self.repo.get_job(app.job_id)
            ws.append(
                [
                    job.company_name if job else "",
                    job.title if job else "",
                    app.status,
                    app.stage,
                    _fmt_dt(app.submitted_at),
                    _fmt_dt(app.responded_at),
                    app.notes or "",
                ]
            )
            if job:
                self._color_row(ws, ws.max_row, job.state, 7)
        _finalize(ws)

    def _build_companies(self, ws: Worksheet) -> None:
        headers = ["Name", "Website", "Industry", "Score", "Open Jobs", "Notes"]
        self._write_header(ws, headers)
        counts = self.repo.count_jobs_by_state()  # not per-company; simple total below
        for company in self.repo.session.query(Company).order_by(Company.name).all():
            open_jobs = len(company.jobs)
            ws.append(
                [
                    company.name,
                    company.website or "",
                    company.industry or "",
                    round(company.score, 3) if company.score is not None else "",
                    open_jobs,
                    company.notes or "",
                ]
            )
        _finalize(ws)
        _ = counts

    def _build_documents(self, ws: Worksheet) -> None:
        headers = ["Job", "Company", "Kind", "Format", "Version", "Path", "Created"]
        self._write_header(ws, headers)
        for doc in self.repo.session.query(Document).order_by(Document.created_at.desc()).all():
            job = self.repo.get_job(doc.job_id) if doc.job_id else None
            ws.append(
                [
                    job.title if job else "",
                    job.company_name if job else "",
                    doc.kind,
                    doc.fmt,
                    doc.version,
                    doc.path,
                    _fmt_dt(doc.created_at),
                ]
            )
        _finalize(ws)

    def _build_statistics(self, ws: Worksheet) -> None:
        from job_agent.analytics.analytics import Analytics

        stats = Analytics(self.repo).compute()
        ws.append(["Metric", "Value"])
        self._style_header_row(ws, 2)
        for key, value in stats.as_rows():
            ws.append([key, value])
        _finalize(ws, autofilter=False)

    # -- helpers ------------------------------------------------------------
    def _write_header(self, ws: Worksheet, headers: list[str]) -> None:
        ws.append(headers)
        self._style_header_row(ws, len(headers))

    def _style_header_row(self, ws: Worksheet, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")

    def _color_row(self, ws: Worksheet, row: int, state: str, ncols: int) -> None:
        color = _STATE_FILL.get(state)
        if not color:
            return
        fill = PatternFill("solid", fgColor=color)
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = fill


def _finalize(ws: Worksheet, *, autofilter: bool = True) -> None:
    ws.freeze_panes = "A2"
    if autofilter and ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
    _autosize(ws)


def _autosize(ws: Worksheet) -> None:
    for col in ws.columns:
        length = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 60)


def _fmt_dt(value: Any) -> str:
    if value is None:
        return ""
    try:
        return value.strftime("%Y-%m-%d %H:%M")
    except AttributeError:
        return str(value)
